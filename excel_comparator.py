import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from collections import defaultdict
import subprocess
from pathlib import Path
from tkinter import ttk
import tkinter.messagebox as messagebox
from tkinter import Button
from tkinter import PhotoImage
from datetime import date
import webbrowser
import logging as log
import colorama
import openpyxl


colorama.init()

                                                ### FENETRE PRINCIPALE ###

# Configuration initiale du logging
log.basicConfig(level=log.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Créer la fenêtre principale
print('-----------------------')
print(colorama.Fore.YELLOW + "Lancement de l\'application..." + colorama.Style.RESET_ALL)
print ('Lançement de l\'interface graphique...')
root = tk.Tk()
root.title('Traitement des fichiers Excel')
print('Ouverture fenêtre principale...')
print('-----------------------')
#icone fenêtres
icone = tk.PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\logo_jr2.png')
# Définir l'icône pour la fenêtre principale
root.tk.call('wm', 'iconphoto', root._w, icone)

# Définir la taille de la fenêtre
root.geometry("850x680")
log.info("Application démarrée")


                                                    ### FONCTIONS ###

# Fonction pour changer l'icône des nouvelles fenêtres
def changer_icone_fenetre(fenetre):
    log.debug("Personalisation de la fenêtre")
    fenetre.iconphoto(True, icone)


# Fonction pour traiter le fichier des livraisons
def traiter_chauffeur():
    log.info(colorama.Fore.YELLOW +"Traitement du fichier des livraisons"+ colorama.Style.RESET_ALL)
    fichier_chauffeur = filedialog.askopenfilename(filetypes=[('Fichiers Excel', '*.xlsx')])

    # Valider le fichier avant de continuer
    if not valider_fichier_chauffeur(fichier_chauffeur):
        log.error(colorama.Fore.RED +"Le fichier des livraisons sélectionné n'est pas au bon format ou ne contient pas les données attendues."+ colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #100", "Le fichier sélectionné n'est pas au bon format ou ne contient pas les données attendues.")
        return

    df_livraisons = pd.read_excel(fichier_chauffeur)

    # Extraire les colonnes des vidanges (de D à N)
    log.debug("Extraction des colonnes des vidanges")
    colonnes_vidanges = df_livraisons.columns[3:14]  # colonnes D à N 
    # Créer un dictionnaire pour stocker les clients, leurs vidanges et les quantités correspondantes
    log.debug("Création d'un dictionnaire pour stocker les clients, leurs vidanges et les quantités correspondantes")
    clients_vidanges = defaultdict(dict)

    print('--------------------')
    print(f"Lecture du fichier '{fichier_chauffeur}' en cours...")
    # Afficher le résultat dans la fenêtre
    result_label.config(text=f"Lecture du fichier '{fichier_chauffeur}' en cours...")

    # Parcourir chaque ligne du DataFrame des livraisons
    for index, row in df_livraisons.iterrows():
        log.debug(f"Lecture de la ligne {index}")
        client = row['Customer Name']
        for vidange in colonnes_vidanges:
            quantite = row[vidange]
            if pd.notna(quantite):
                if vidange not in clients_vidanges[client]:
                    clients_vidanges[client][vidange] = quantite
                else:
                    clients_vidanges[client][vidange] += quantite

    # Fonction pour mettre à jour la barre de progression
    def update_progress(progress):
        progress_bar.step(progress)
        root.update_idletasks()

    nb_etapes = 5

    # Mettre à jour la barre de progression progressivement
    for i in range(1, nb_etapes + 1):
        log.debug(f"Etape de traitement {i} sur {nb_etapes}")
        root.after(i * 100, update_progress, 49 // nb_etapes)


    # Créer une liste de listes pour le tableau
    table_data = []
    for client, vidanges in clients_vidanges.items():
        row = [client] + [vidanges.get(vidange, 0) for vidange in colonnes_vidanges]
        log.debug(f"Création de la ligne {row}")
        table_data.append(row)

    # Afficher le tableau
    headers = ['Client'] + list(colonnes_vidanges)
    log.debug(f"Affichage du tableau")

    # Convertir la liste de listes en DataFrame pandas
    df_output = pd.DataFrame(table_data, columns=headers)

    # Enregistrer le DataFrame dans un fichier Excel dans le dossier des téléchargements
    log.debug(f"Enregistrement du fichier Excel dans le dossier des téléchargements")
    nom_fichier_excel = os.path.join(os.path.expanduser('~'), 'Downloads', f'df_chauffeur_{date.today()}.xlsx')
    df_output.to_excel(nom_fichier_excel, index=False)

    print(colorama.Fore.BLUE +f"Le fichier Excel concernant '{fichier_chauffeur}' a été enregistré avec succès dans le dossier téléchargements sous le nom:", nom_fichier_excel+ colorama.Style.RESET_ALL)
    # Afficher le résultat dans la fenêtre
    result_label.config(text=f"Le fichier Excel a été enregistré avec succès\n")
    messagebox.showinfo("Prêt", "Le fichier a bien été enregistré !\n\nVous pouvez maintenant traiter le fichier des vidanges.")


# Fonction pour traiter le fichier des vidanges
def traiter_descartes():
    log.info(colorama.Fore.YELLOW + "Traitement du fichier des vidanges" + colorama.Style.RESET_ALL)
    fichier_descartes = filedialog.askopenfilename(filetypes=[('Fichiers Excel', '*.xlsx')])

    # Valider le fichier avant de continuer
    if not valider_fichier_descartes(fichier_descartes):
        log.error(colorama.Fore.RED + "Le fichier des vidanges sélectionné n'est pas au bon format ou ne contient pas les données attendues." + colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #100", "Le fichier sélectionné n'est pas au bon format ou ne contient pas les données attendues.")
        return

    df_vidanges = pd.read_excel(fichier_descartes)

    # Remplir les valeurs manquantes dans la colonne "Client" pour associer chaque article au client approprié
    df_vidanges['Client'].fillna(method='ffill', inplace=True)

    # Prendre la valeur absolue des montants pour les considérer comme positifs
    df_vidanges['Lignes de la commande/Quantité'] = df_vidanges['Lignes de la commande/Quantité'].abs()

    # Pivoter les données pour obtenir la nouvelle structure avec la somme des quantités facturées
    df_descartes = df_vidanges.pivot_table(index='Client', columns='Lignes de la commande/Article', values='Lignes de la commande/Quantité', aggfunc='sum')

    # Remplacer les valeurs NaN par 0
    df_descartes.fillna(0, inplace=True)

    # Exporter le tableau en fichier Excel
    fichier_sortie = os.path.join(os.path.expanduser('~'), 'Downloads', f'df_descartes_{date.today()}.xlsx')
    df_descartes.to_excel(fichier_sortie)

    def update_progress(progress):
        progress_bar.step(progress)
        root.update_idletasks()

    nb_etapes = 5

    # Mettre à jour la barre de progression progressivement
    for i in range(1, nb_etapes + 1):
        log.debug(f"Etape de traitement {i} sur {nb_etapes}")
        root.after(i * 100, update_progress, 50 // nb_etapes)
    print(colorama.Fore.BLUE + f"Le fichier Excel '{fichier_sortie}' a été créé avec succès dans le dossier téléchargements." + colorama.Style.RESET_ALL)
    result_label.config(text=f"Le fichier Excel a été enregistré avec succès \n")
    messagebox.showinfo("Prêt", "Le fichier a bien été enregistré !\n\nVous pouvez maintenant comparer les deux fichiers générés.")


# Fonctions de validation pour les fichiers

def valider_fichier_descartes(fichier):
    # Vérifier que le fichier est au format xlsx
    print('----------')
    print(colorama.Fore.GREEN +"Vérification du format du fichier en cours..."+ colorama.Style.RESET_ALL)
    if not fichier.lower().endswith('.xlsx'):
        log.critical(colorama.Fore.RED +"Le fichier sélectionné n'est pas au bon format."+ colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #101", "Le fichier sélectionné n'est pas au bon format.")
        return False

    # Vérifier que le fichier contient les colonnes attendues
    print(colorama.Fore.GREEN +'Vérification du contenu du fichier en cours...'+ colorama.Style.RESET_ALL)
    colonnes_attendues = ['Client', 'Lignes de la commande/Article', 'Lignes de la commande/Quantité']
    df = pd.read_excel(fichier)
    colonnes_fichier = df.columns.tolist()

    if not all(colonne in colonnes_fichier for colonne in colonnes_attendues):
        messagebox.showerror("Erreur #102", "Le fichier sélectionné ne contient pas les données attendues.")
        log.critical(colorama.Fore.RED +"Le fichier sélectionné ne contient pas les données attendues."+ colorama.Style.RESET_ALL)
        return False

    return True


def valider_fichier_chauffeur(fichier):
    # Vérifier que le fichier est au format xlsx
    print('----------')
    print(colorama.Fore.GREEN +"Vérification du format du fichier en cours..."+ colorama.Style.RESET_ALL)
    if not fichier.lower().endswith('.xlsx'):
        log.critical(colorama.Fore.RED +"Le fichier sélectionné n'est pas au bon format."+ colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #101", "Le fichier sélectionné n'est pas au bon format.")
        return False

    # Vérifier que le fichier contient les colonnes attendues
    print(colorama.Fore.GREEN +'Vérification du contenu du fichier en cours...'+ colorama.Style.RESET_ALL)
    colonnes_attendues = ['Customer Name', 'Palette Euro NEW', 'Caisses vertes', 'VID-T', 'VID-S', 'Vidange F', 'FRIGO BOX', 'Palette Truval', 'Palette banane', 'Palette Plastique', 'Palette Pool' ]
    df = pd.read_excel(fichier)
    colonnes_fichier = df.columns.tolist()

    if not all(colonne in colonnes_fichier for colonne in colonnes_attendues):
        log.critical(colorama.Fore.RED +"Le fichier sélectionné ne contient pas les données attendues."+ colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #102", "Le fichier sélectionné ne contient pas les données attendues.")
        return False

    return True


def charger_df_descartes():
    fichier_descartes = os.path.join(os.path.expanduser('~'), 'Downloads', f'df_descartes_{date.today()}.xlsx')
    if os.path.exists(fichier_descartes):
        df_descartes = pd.read_excel(fichier_descartes)
        return df_descartes
    else:
        return None

def charger_df_chauffeur():
    fichier_chauffeur = os.path.join(os.path.expanduser('~'), 'Downloads', f'df_chauffeur_{date.today()}.xlsx')
    if os.path.exists(fichier_chauffeur):
        df_chauffeur = pd.read_excel(fichier_chauffeur)
        return df_chauffeur
    else:
        return None

def comparer_fichiers():
    log.info(colorama.Fore.YELLOW + "Comparaison des fichiers générés" + colorama.Style.RESET_ALL)

    # Charger les DataFrames à partir des fichiers Excel
    df_descartes = charger_df_descartes()
    df_chauffeur = charger_df_chauffeur()

    log.debug("Vérification de l'existance des fichiers")
    if df_descartes is None or df_chauffeur is None:
        log.critical(colorama.Fore.RED + "Il n'y a pas de fichier à comparer et/ou il en manque un !" + colorama.Style.RESET_ALL)
        print(colorama.Fore.RED + "Les fichiers 'df_descartes' et/ou 'df_chauffeur' n'existent pas." + colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #200", "Il n'y a pas de fichier à comparer et/ou il en manque un !\nVeuillez traiter les fichiers avant de les comparer")
        return

    log.debug("Vérification du contenu des fichiers")
    # Vérifier que les fichiers contiennent des données
    if df_descartes.empty or df_chauffeur.empty:
        log.critical(colorama.Fore.RED + "Les fichiers ne contiennent pas de données." + colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #103", "Les fichiers ne contiennent pas de données.")
        return

    log.debug("Mise en relation des colonnes des deux fichiers")
    # Correspondance des colonnes entre les deux fichiers
    correspondance_colonnes = {
        'Client': 'Client',
        'PALETTE EU 11': 'Palette Euro NEW',
        'PALETTE EU 9' : 'Palette Euro NEW',
        'EPS 246': 'Caisses vertes',
        'EPS - T': 'VID-T',
        'EPS - M': 'VID-S',
        'VID F': 'Vidange F',
        'FRIGOBOX': 'FRIGO BOX',
        'PALETTE TRUVAL': 'Palette Truval',
        'PALETTE BANANE': 'Palette banane',
        'PALETTE PLASTIQUE': 'Palette Plastique',
        'PALETTE POOL': 'Palette Pool'
    }

    # Utiliser les colonnes "Client" comme index pour les deux DataFrames
    df_descartes.set_index('Client', inplace=True)
    df_chauffeur.set_index('Client', inplace=True)

    log.debug("Création d'un DataFrame pour stocker les écarts")
    # Créer un DataFrame df_ecarts pour stocker les écarts entre les quantités de vidanges
    df_ecarts = df_descartes.copy()

    log.info("Calcul des écarts")
    # Calculer les écarts pour les quantités de vidanges
    for colonne_descartes, colonne_chauffeur in correspondance_colonnes.items():
        if colonne_descartes in df_descartes.columns and colonne_chauffeur in df_chauffeur.columns:
            descartes_values = pd.to_numeric(df_descartes[colonne_descartes], errors='coerce')
            chauffeur_values = pd.to_numeric(df_chauffeur[colonne_chauffeur], errors='coerce')
            ecarts = descartes_values - chauffeur_values
            df_ecarts[colonne_descartes] = ecarts

    log.debug("Suppression des clients sans écarts")
    # Supprimer les clients sans écart du DataFrame "df_ecarts"
    df_ecarts = df_ecarts[df_ecarts.iloc[:, 1:].ne(0).any(axis=1)]
    
    # Vérifier si le DataFrame est vide
    if df_ecarts.empty:
        messagebox.showinfo("Aucun écart", "Il n'y a pas d'écart concernant les clients à cette date !")
        log.info(colorama.Fore.GREEN + "Il n'y a pas d'écarts concernant les clients à cette date !" + colorama.Style.RESET_ALL)
        return
    

    log.debug("Enregistrement du fichier Excel dans le dossier des téléchargements")
    # Exporter les écarts en fichier Excel dans le dossier des téléchargements
    fichier_sortie = os.path.join(os.path.expanduser('~'), 'Downloads', f'ecarts_{date.today()}.xlsx')
    df_ecarts.reset_index(inplace=True)  # Réinitialiser l'index pour inclure à nouveau la colonne "Client" dans le fichier Excel
    df_ecarts.to_excel(fichier_sortie, index=False)

    # Ouvrir le fichier Excel avec openpyxl
    wb = openpyxl.load_workbook(fichier_sortie)
    ws = wb.active

    log.debug("Mise en forme du fichier Excel")
    # Parcourir les cellules du fichier Excel et appliquer le format en rouge pour les valeurs négatives et en vert pour les valeurs positives
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        for col_idx, cell in enumerate(row[1:], start=2):  # Commencer à partir de la deuxième colonne (la première colonne est le client)
            if isinstance(cell, (int, float)) and cell < 0:
                ws.cell(row=idx, column=col_idx).font = openpyxl.styles.Font(color='FF0000')  # Rouge pour les valeurs négatives
            elif isinstance(cell, (int, float)) and cell > 0:
                ws.cell(row=idx, column=col_idx).font = openpyxl.styles.Font(color='00FF00')  # Vert pour les valeurs positives

    # Sauvegarder le fichier Excel modifié
    wb.save(fichier_sortie)

    print(colorama.Fore.BLUE + 'La comparaison est terminée !' + colorama.Style.RESET_ALL)
    print(colorama.Fore.BLUE + f"Le fichier Excel '{fichier_sortie}' a été créé avec succès." + colorama.Style.RESET_ALL)
    messagebox.showinfo("Validation", "Les fichiers ont bien été comparés !\n\nTous les fichiers générés sont disponibles dans le dossier des téléchargements.")
    messagebox.showwarning("Présence d'écarts", "Des écarts sont présents concernant les clients à cette date !\n\nVeuillez ouvrir le fichier Excel des écarts pour les consulter.")
    result_label.config(text="La comparaison est terminée !\nLe fichier Excel a été enregistré avec succès\n")




# Fonction pour ouvrir le dernier fichier généré
def ouvrir_dernier_fichier():
    log.info(colorama.Fore.YELLOW +"Ouverture du dernier fichier généré"+ colorama.Style.RESET_ALL)
    progress_bar.step(100)
    dossier_telechargements = Path.home() / 'Downloads'

    fichier_ecarts = f'ecarts_{date.today()}.xlsx'

    chemin_fichier_genere = dossier_telechargements / fichier_ecarts

    if chemin_fichier_genere.exists():
        log.debug(f"Ouverture du fichier '{fichier_ecarts}'")
        progress_bar.step(100)
        os.startfile(str(chemin_fichier_genere))
        print(colorama.Fore.BLUE +f"Le dernier fichier généré '{fichier_ecarts}' a été ouvert."+ colorama.Style.RESET_ALL)
        result_label.config(text=f"Le dernier fichier généré a été ouvert : {chemin_fichier_genere}")
    else:
        log.critical(colorama.Fore.RED +"Le fichier des comparaisons n'existe pas."+ colorama.Style.RESET_ALL)
        print(colorama.Fore.RED +f"Le fichier '{fichier_ecarts}' n'existe pas."+ colorama.Style.RESET_ALL)
        messagebox.showerror("Erreur #300", "Le fichier des comparaisons n'existe pas.\nVeuillez traiter les fichiers avant de les comparer")


# Fonction pour ouvrir le dossier des téléchargements
def ouvrir_dossier_telechargements():
    log.info(colorama.Fore.YELLOW +"Ouverture du dossier des téléchargements"+ colorama.Style.RESET_ALL)
    progress_bar.step(100)
    dossier_telechargements = os.path.join(os.path.expanduser('~'), 'Downloads')
    subprocess.Popen(f'explorer "{dossier_telechargements}"')
    print(colorama.Fore.BLUE +f"Le dossier des téléchargements a été ouvert : {dossier_telechargements}"+ colorama.Style.RESET_ALL)
    result_label.config(text=f"Le dossier des téléchargements a été ouvert : {dossier_telechargements}")


# Fonction pour quitter la fenêtre
def quitter_fenetre():
    root.quit()
    print('-----------------------')
    log.info(colorama.Fore.YELLOW +"Fermeture de l'application"+ colorama.Style.RESET_ALL)
    print(colorama.Fore.YELLOW +'Fermeture de l\'application...'+ colorama.Style.RESET_ALL)
    quit_label = tk.Label(root, text="Fermeture de l'application...")
    quit_label.pack()


def afficher_aide():
    message_aide = """
    RUBRIQUE D'AIDE : \n\n
    Tous les fichiers générés sont disponibles dans le dossier des téléchargements. \n\nLeur nom contient la date du jour. \n\n
    IMPORTANT : \n
    - Seuls les fichiers Excel en .xlsx sont acceptés.\n
    - Le fichier des livreurs doit avoir ces colonnes : 'Customer Name', 'Palette Euro NEW', 'Caisses vertes', 'VID-T', 'VID-S', 'Vidange F', 'FRIGO BOX', 'Palette Truval', 'Palette banane', 'Palette Plastique', 'Palette Pool' et les noms ne doivent pas être changés. Peu engendrer des erreurs.\n
    - Le fichier Descartes doit avoir ces colonnes : 'Client', 'Lignes de la commande/Article', 'Lignes de la commande/Quantité' et les noms ne doivent pas être changés. Peu engendrer des erreurs.\n
    - Bien attendre la fenêtre pop-up avant de cliquer sur le bouton suivant.
    - Il est conseillé de suivre l'ordre des boutons pour éviter les erreurs. \n\n
    En cas d'erreur : \n
    - Lire le message d'erreur, une rubrique traitant les erreurs est disponible dans l'onglet 'Aide'.
    - Vérifiez le format des fichiers Excel (.xlsx).
    - Vérifiez que les fichiers Excel sont correctement structuré et ne contiennent pas d'erreurs.\n
    Si toutes ces vérifications sont correctes, veuillez réessayer les opérations dans l'ordre.
    """
    messagebox.showinfo("Aide", message_aide)
    log.debug(colorama.Fore.GREEN +"Affichage de l'aide"+ colorama.Style.RESET_ALL)

def convert_to_xlsx():
    message_xlsx = """
    Pour convertir un fichier Excel en .xlsx : \n
    - Ouvrez le fichier Excel à convertir.\n
    - Cliquez sur 'Fichier' puis 'Enregistrer sous'.\n
    - Dans la liste déroulante 'Type', sélectionnez 'Classeur Excel (*.xlsx)'.\n
    
    Le fichier est maintenant converti en .xlsx.
    """
    messagebox.showinfo("Conversion en .xlsx", message_xlsx)
    log.debug(colorama.Fore.GREEN +"Affichage de la convertion"+ colorama.Style.RESET_ALL)
    

def afficher_instructions_boutons():
    message = f"""
    Instructions d'utilisation :
    - Bouton 'Traiter le fichier des livraisons': Cliquez sur ce bouton pour traiter le fichier des livraisons. Une fois le traitement terminé, le fichier généré sera enregistré dans le dossier des téléchargements.

    - Bouton 'Traiter le fichier des vidanges': Cliquez sur ce bouton pour traiter le fichier des vidanges. Une fois le traitement terminé, le fichier généré sera enregistré dans le dossier des téléchargements.

    - Bouton 'Comparer les fichiers générés': Cliquez sur ce bouton pour comparer les deux fichiers générés. Une fois la comparaison terminée, le résultat sera enregistré dans un nouveau fichier Excel dans le dossier des téléchargements.

    - Bouton 'Ouvrir le dernier fichier généré': Cliquez sur ce bouton pour ouvrir le dernier fichier Excel généré. Assurez-vous que le fichier existe dans le dossier des téléchargements.

    - Bouton 'Ouvrir le dossier des téléchargements': Cliquez sur ce bouton pour ouvrir le dossier des téléchargements où tous les fichiers générés sont enregistrés.

    - Bouton 'Quitter l'application': Cliquez sur ce bouton pour fermer l'application.

    Note : Suivez l'ordre des boutons pour éviter les erreurs lors du traitement des fichiers.
    """
    messagebox.showinfo("Instructions d'utilisation", message)
    log.debug(colorama.Fore.YELLOW +"Affichage des instructions d'utilisation des boutons"+ colorama.Style.RESET_ALL)


def err_100():
    message_100 = """
    Erreur 100 : Le fichier sélectionné n'est pas au bon format ou ne contient pas les données attendues.
    Pour régler cette erreur :
    - Assurez-vous que le fichier est au format Excel (.xlsx).
    - Vérifiez que le fichier contient les colonnes attendues : 'Client', 'Lignes de la commande/Article', 'Lignes de la commande/Quantité facturée'.
    """
    messagebox.showinfo("Erreur 100", message_100)
    log.debug(colorama.Fore.GREEN +"Affichage de l'erreur 100"+ colorama.Style.RESET_ALL)

def err_101():
    message_101 = """
    Erreur 101 : Le fichier sélectionné n'est pas au bon format.
    Pour régler cette erreur :
    - Assurez-vous que le fichier est au format Excel (.xlsx).
    """
    messagebox.showinfo("Erreur 101", message_101)
    log.debug(colorama.Fore.GREEN +"Affichage de l'erreur 101"+ colorama.Style.RESET_ALL)

def err_102():
    message_102 = """
    Erreur 102 : Le fichier sélectionné ne contient pas les données attendues.
    Pour régler cette erreur :
    - Vérifiez que le fichier contient les colonnes attendues : 'Client', 'Lignes de la commande/Article', 'Lignes de la commande/Quantité facturée'.
    """
    messagebox.showinfo("Erreur 102", message_102)
    log.debug(colorama.Fore.GREEN +"Affichage de l'erreur 102"+ colorama.Style.RESET_ALL)

def err_103():
    message_103 = """
    Erreur 103 : Les fichiers ne contiennent pas de données.
    Ce message indique que le fichier est vide. Assurez-vous de choisir le bon fichier.
    """
    messagebox.showinfo("Erreur 103", message_103)
    log.debug(colorama.Fore.GREEN +"Affichage de l'erreur 103"+ colorama.Style.RESET_ALL)

def err_200():
    message_200 = """
    Erreur 200 : Il n'y a pas de fichier à comparer et/ou il en manque un !
    Ce message indique qu'il manque un fichier à comparer. Assurez-vous de traiter les fichiers avant de les comparer.
    """
    messagebox.showinfo("Erreur 200", message_200)
    log.debug(colorama.Fore.GREEN +"Affichage de l'erreur 200"+ colorama.Style.RESET_ALL)

def err_300():
    message_300 = """
    Erreur 300 : Le fichier des comparaisons n'existe pas.
    Ce message indique que la comparaison n'a pas été effectuée.
    """
    messagebox.showinfo("Erreur 300", message_300)
    log.debug(colorama.Fore.GREEN +"Affichage de l'erreur 300"+ colorama.Style.RESET_ALL)


def type_erreur():
    types_errors_window = tk.Toplevel(root)
    types_errors_window.title("Types d'erreurs")
    types_errors_window.geometry("700x450")  # Définir la taille de la fenêtre
    log.debug(colorama.Fore.YELLOW +"Ouverture de la fenêtre des types d'erreurs"+ colorama.Style.RESET_ALL)

    # Ajouter les boutons d'erreurs à la fenêtre pop-up
    tk.Button(types_errors_window, text="Erreur 100 : Le fichier sélectionné n'est pas au bon format ou ne contient pas les données attendues", command=err_100, image=ico_error, compound='left', width=410,wraplength=400).pack(padx=20, pady=10)
    tk.Button(types_errors_window, text="Erreur 101 : Le fichier sélectionné n'est pas au bon format", command=err_101, image=ico_error, compound='left', width=410,wraplength=400).pack(padx=20, pady=10)
    tk.Button(types_errors_window, text="Erreur 102 : Le fichier sélectionné ne contient pas les données attendues", command=err_102, image=ico_error, compound='left', width=410,wraplength=400).pack(padx=20, pady=10)
    tk.Button(types_errors_window, text="Erreur 103 : Les fichiers ne contiennent pas de données", command=err_103, image=ico_error, compound='left', width=410,wraplength=400).pack(padx=20, pady=10)
    tk.Button(types_errors_window, text="Erreur 200 : Il n'y a pas de fichier à comparer et/ou il en manque un !\nVeuillez traiter les fichiers avant de les comparer", command=err_200, image=ico_error, compound='left', width=410,wraplength=400).pack(padx=20, pady=10)
    tk.Button(types_errors_window, text="Erreur 300 : Le fichier des comparaisons n'existe pas.\nVeuillez traiter les fichiers avant de les comparer", command=err_300, image=ico_error, compound='left', width=410,wraplength=400).pack(padx=20, pady=10)

    # Ajouter le bouton "OK" pour fermer la fenêtre
    btn_ok = tk.Button(types_errors_window, text="OK", command=types_errors_window.destroy, image=ico_ok, compound='left')
    btn_ok.pack(pady=10)


def ouvrir_github():
    # Ouvrir le lien GitHub dans le navigateur web par défaut
    webbrowser.open("https://github.com/Nicolo41/excel_comparator")
    log.debug("Ouverture du lien GitHub")

def afficher_fct() :
    message = """
    L'application suivante sert à traiter des fichiers Excel. \nCette interface utilisateur ainsi que les fonctions ont été développées en Python. \n\n
    L'application permet de comparer deux fichiers Excel et de voir s'il y a des différences entre les deux. \n\n

    PROCESUS : \n
    - L'utilisateur doit sélectionner un fichier Excel contenant les données des livraisons. \n Le script va lire le fichier et créer un nouveau fichier Excel contenant les données des livraisons par client. \n\n
    - L'utilisateur doit sélectionner un fichier Excel contenant les données des vidanges. \n Le script va lire le fichier et créer un nouveau fichier Excel contenant les données des vidanges par client. \n\n
    - Lorsque l'utilisateur appuie sur le bouton 'Comparer les fichiers générés', le script va comparer les deux fichiers générés et créer un nouveau fichier Excel contenant les différences entre les deux. \n\n

    Le code est fait pour que chaque fichier généré soit enregistré dans le dossier des téléchargements à la date de la création. \n\n
    Lien vers le code source :  \n\n
    """
    # messagebox.showinfo("Fonctionnement de l'application", message)
    types_errors_window = tk.Toplevel(root)
    types_errors_window.title("Fonctionnement de l'application")

    # Afficher le message dans la fenêtre pop-up
    tk.Label(types_errors_window, text=message, wraplength=600).pack(padx=20, pady=10)

    # Ajouter un bouton pour ouvrir le lien GitHub
    btn_github = tk.Button(types_errors_window, text="GitHub", command=ouvrir_github, image=ico_git, compound='left')
    btn_github.pack(pady=20, padx=20)
    # Ajouter le bouton "OK" pour fermer la fenêtre
    btn_ok = tk.Button(types_errors_window, text="OK", command=types_errors_window.destroy, image=ico_ok, compound='left')
    btn_ok.pack(pady=10)

    log.debug(colorama.Fore.YELLOW +"Affichage du fonctionnement de l'application"+ colorama.Style.RESET_ALL)

                                              ### ICONES ET IMAGES ###

# Charger les icônes au format .png avec tkinter

log.debug("Chargement des icônes")
ico_excel = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\excel2.png')
ico_compare = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\compare2.png')
ico_exit = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\exit2.png')
ico_fichier = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\fichier2.png')
ico_folder = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\folder2.png')
ico_error = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\error2.png')
ico_git = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\git2.png')
ico_ok = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\ok.png')
#background
bg = PhotoImage(file='C:\\Users\\nicol\\Documents\\Documents\\Code\\Python\\Vidanges\\ecarts_vidanges\\img\\logo_jr.png')


                                            ### WIDGETS, BOUTONS ET MENUS ###


# Créer un widget Label pour afficher l'image en arrière-plan
background_label = tk.Label(root, image=bg)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

# Ajouter d'autres widgets sur la fenêtre
label1 = tk.Label(root, text="Bienvenue !", font=('Arial', 20))
label1.pack(pady=20)

# Créer un menu
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# Créer un sous-menu "Aide"
menu_aide = tk.Menu(menu_bar, tearoff=False)
menu_bar.add_cascade(label='Aide', menu=menu_aide)

# Ajouter un élément dans le sous-menu "Aide" pour afficher l'aide
menu_aide.add_command(label='Général', command=afficher_aide)
menu_aide.add_command(label='Conversion en .xlsx', command=convert_to_xlsx)
menu_aide.add_command(label='Fonctionnement de l\'application', command=afficher_fct)
menu_aide.add_command(label='Afficher les instructions d\'utilisation des boutons', command=afficher_instructions_boutons)
menu_aide.add_command(label='Types d\'erreurs', command=type_erreur)


# Ajouter un widget Label pour afficher les résultats
result_label = tk.Label(root, text="")
result_label.pack(padx=20, pady=10)


result_label.config(text="Rendez-vous dans la rubrique 'Aide' en haut à gauche pour plus d'informations\nUne pop-up vous indiquera quand vous pourrez cliquer sur le bouton suivant.\n\nVous disposez d'une fenêtre avec les logs qui afficherons les erreurs", font=('Arial', 10))

# Créer des boutons pour les différentes opérations
btn_traiter_livraisons = tk.Button(root, text='1. Traiter le fichier des chauffeurs', command=traiter_chauffeur, image = ico_excel, compound='left', font=('Arial', 10))
btn_traiter_livraisons.pack(padx=20, pady=10)

btn_traiter_vidanges = tk.Button(root, text='2. Traiter le fichier de Descartes', command=traiter_descartes, image = ico_excel, compound='left', font=('Arial', 10))
btn_traiter_vidanges.pack(padx=20, pady=10)

btn_comparer_fichiers = tk.Button(root, text='3. Comparer les fichiers traités', command=comparer_fichiers, image = ico_compare, compound='left', font=('Arial', 10))
btn_comparer_fichiers.pack(padx=20, pady=10)

# Créer un bouton pour ouvrir le dernier fichier généré
btn_ouvrir_fichier = tk.Button(root, text='4. Ouvrir le fichier des écarts', command=ouvrir_dernier_fichier, image = ico_fichier, compound='left', font=('Arial', 10))
btn_ouvrir_fichier.pack(padx=20, pady=10)

# Créer un bouton pour ouvrir le dossier des téléchargements
btn_ouvrir_dossier = tk.Button(root, text='Ouvrir le dossier des téléchargements', command=ouvrir_dossier_telechargements, image = ico_folder, compound='left', font=('Arial', 10))
btn_ouvrir_dossier.pack(padx=20, pady=10)

# Ajouter une barre de progression
progress_bar = ttk.Progressbar(root, mode='determinate', maximum=100)
progress_bar.pack(fill='x', padx=20, pady=10)

# Créer un bouton pour quitter la fenêtre
btn_quitter = Button(root, text='Quitter l\'application', foreground="red", command=quitter_fenetre, image = ico_exit, compound='left', font=('Arial', 10))
btn_quitter.pack(padx=20, pady=10)


# Changer l'icône des nouvelles fenêtres
changer_icone_fenetre(root)

# Lancer l'interface graphique
root.mainloop()